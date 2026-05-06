"""Google Ads Keyword Overlap Detector v1.0

Detects duplicate / overlapping keywords across ad groups within a single account
that bid against each other (cannibalization).

Usage:
    python detect_overlap.py <keywords_csv> [output_path]

Input: Google Ads keywords export (CSV from Editor or UI). Required columns:
  - Campaign / Campagne
  - Ad group / Advertentiegroep
  - Keyword / Zoekwoord
  - Match type / Type overeenkomst
  - Status / Status

Conflict types detected:
  1. EXACT_DUPLICATE   — same keyword + same match type in 2+ ad groups
  2. MATCH_TYPE_OVERLAP — same keyword text, different match types in 2+ AGs
  3. PHRASE_CONTAINS   — Broad keyword X also appears as fully contained sub-phrase
  4. CLOSE_VARIANT     — singular/plural / minor stem variations across AGs

Annotations:
  - "Probably geo-segmented" — TRUE if campaign names end with country/region codes
    suggesting intentional geo-split (NL/DE/FR/UK/EU/etc)
  - "All paused" — TRUE if all instances are already paused (informational only)

Output: Excel with 2 sheets.
  Sheet 1 "Overlap Issues" — full conflict list, color-coded by severity
  Sheet 2 "Suggested Pauses" — Editor-importable paste-ready format
"""
import csv
import os
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ── Color fills ──────────────────────────────────────────────────────────────
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")  # high severity
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # medium severity
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")  # low severity
FILL_GREY   = PatternFill("solid", fgColor="E7E6E6")  # informational

SEVERITY_FILL = {
    "EXACT_DUPLICATE":     FILL_RED,
    "MATCH_TYPE_OVERLAP":  FILL_YELLOW,
    "PHRASE_CONTAINS":     FILL_YELLOW,
    "CLOSE_VARIANT":       FILL_GREEN,
}


# ── Geo / region tokens for campaign-name pattern detection ──────────────────
# When two conflicting keywords' campaign names differ only in a trailing geo
# token, mark the conflict as "Probably geo-segmented" so the user can quickly
# triage these as intentional segmentation.
_GEO_TOKENS = {
    # Country codes (ISO + common variants)
    "nl", "be", "de", "fr", "uk", "gb", "us", "ca", "it", "es", "pt",
    "au", "nz", "ie", "se", "no", "dk", "fi", "pl", "cz", "sk", "hu",
    "ro", "bg", "gr", "tr", "ru", "ua", "ch", "at", "lu",
    "in", "jp", "cn", "kr", "sg", "hk", "tw", "ph", "id", "my", "th",
    "vn", "br", "mx", "ar", "cl", "co", "pe", "za", "ae", "il", "sa",
    # Regions / market segments
    "eu", "emea", "apac", "americas", "latam", "nordic", "nordics",
    "benelux", "dach", "iberia", "scandinavia", "global", "international",
    # Full country names (English + Dutch)
    "netherlands", "germany", "france", "italy", "spain", "portugal",
    "belgium", "austria", "switzerland", "sweden", "norway", "denmark",
    "finland", "poland", "czech", "ireland", "england", "scotland",
    "nederland", "duitsland", "frankrijk", "spanje", "italie",
    "belgie", "oostenrijk", "zwitserland", "polen",
    # Languages (sometimes used for campaign segmentation)
    "english", "dutch", "german", "french", "italian", "spanish",
    "engels", "duits", "nederlands", "frans", "spaans",
}


# ── Match type normalization ─────────────────────────────────────────────────
# Google Ads exports vary: "Exact match", "Exact", "[exact]", etc.
def _normalize_match_type(mt: str) -> str:
    """Returns one of: EXACT, PHRASE, BROAD, BROAD_MOD, UNKNOWN."""
    if not mt:
        return "UNKNOWN"
    s = mt.lower().strip()
    if s.startswith("exact"):
        return "EXACT"
    if s.startswith("phrase"):
        return "PHRASE"
    if s.startswith("broad match modifier") or s.startswith("modified"):
        return "BROAD_MOD"
    if s.startswith("broad"):
        return "BROAD"
    return "UNKNOWN"


# ── Keyword normalization (for grouping/matching) ────────────────────────────
def _strip_decoration(kw: str) -> str:
    """Strip Google Ads match-type decoration: [exact], "phrase", +broad +mod."""
    s = kw.strip()
    # Remove [brackets] for Exact
    if s.startswith("[") and s.endswith("]"):
        s = s[1:-1]
    # Remove "quotes" for Phrase
    if s.startswith('"') and s.endswith('"'):
        s = s[1:-1]
    # Remove leading + for Broad Match Modifier (legacy)
    s = re.sub(r"\+", "", s)
    return s.strip()


def _strip_accents(s: str) -> str:
    """ASCII-fold accents for normalization. 'café' → 'cafe'."""
    if not s:
        return s
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _normalize_keyword(kw: str) -> str:
    """Lowercase, strip decoration, strip accents, collapse whitespace + hyphens."""
    s = _strip_decoration(kw)
    s = _strip_accents(s.lower())
    s = re.sub(r"[\s\-_]+", " ", s)
    return s.strip()


# ── Close-variant detection (light singular/plural stemming) ─────────────────
def _stem_lite(kw: str) -> str:
    """Very light stemming: drop trailing 's', 'es', 'en' for plurals.

    Not as smart as a real stemmer but catches the common cases:
      transport / transports → transport
      bedrijven / bedrijf → similar root
    Only used for CLOSE_VARIANT detection; conservative on purpose.
    """
    s = _normalize_keyword(kw)
    words = s.split()
    if not words:
        return s
    last = words[-1]
    # English plurals
    if len(last) > 4 and last.endswith("ies"):
        last = last[:-3] + "y"
    elif len(last) > 3 and last.endswith("es"):
        last = last[:-2]
    elif len(last) > 3 and last.endswith("s") and not last.endswith("ss"):
        last = last[:-1]
    # Dutch plurals
    elif len(last) > 4 and last.endswith("en"):
        last = last[:-2]
    words[-1] = last
    return " ".join(words)


# ── Geo-segmentation heuristic ───────────────────────────────────────────────
def _is_geo_segmented(campaign_names: list[str]) -> bool:
    """Return True if all campaign names share a prefix and differ only in a
    trailing geo/region token. Used as 'Probably geo-segmented' annotation.

    Examples:
      ['Transport NL', 'Transport DE']            → True
      ['NL Transport', 'DE Transport']            → True
      ['Smartgyro Search EU', 'Smartgyro Search US']  → True
      ['Brand Campaign', 'Generic Search']        → False
      ['Transport NL', 'Smartgyro NL']            → False (different prefix)
    """
    if len(campaign_names) < 2:
        return False
    parsed = []
    for name in campaign_names:
        tokens = re.split(r"[\s\-_/|]+", name.strip())
        tokens = [t.lower() for t in tokens if t]
        if not tokens:
            return False
        # Find geo token at start, end, or anywhere — pick first match
        geo = None
        non_geo = []
        for t in tokens:
            if t in _GEO_TOKENS and geo is None:
                geo = t
            else:
                non_geo.append(t)
        if geo is None:
            return False
        parsed.append((geo, tuple(non_geo)))

    # All non-geo parts must be identical, and geo tokens must differ
    non_geo_sets = {p[1] for p in parsed}
    geo_set = {p[0] for p in parsed}
    return len(non_geo_sets) == 1 and len(geo_set) > 1


# ── Data classes ─────────────────────────────────────────────────────────────
@dataclass
class KeywordRow:
    raw_keyword: str        # as written in CSV (with brackets/quotes)
    keyword: str            # decoration-stripped
    norm_keyword: str       # normalized for grouping
    match_type: str         # normalized: EXACT/PHRASE/BROAD/BROAD_MOD
    raw_match_type: str     # original string from CSV
    campaign: str
    ad_group: str
    status: str
    final_url: str = ""
    row_index: int = 0      # original CSV row number for traceability


@dataclass
class Conflict:
    conflict_type: str      # EXACT_DUPLICATE / MATCH_TYPE_OVERLAP / PHRASE_CONTAINS / CLOSE_VARIANT
    keywords_involved: list[KeywordRow] = field(default_factory=list)
    geo_segmented: bool = False
    all_paused: bool = False

    @property
    def severity_rank(self) -> int:
        """Lower = more severe (for sorting)."""
        ranks = {
            "EXACT_DUPLICATE": 0,
            "MATCH_TYPE_OVERLAP": 1,
            "PHRASE_CONTAINS": 2,
            "CLOSE_VARIANT": 3,
        }
        return ranks.get(self.conflict_type, 99)

    @property
    def display_keyword(self) -> str:
        """Best-effort representative keyword for display."""
        if not self.keywords_involved:
            return ""
        return self.keywords_involved[0].keyword


# ═════════════════════════════════════════════════════════════════════════════
# CSV reading
# ═════════════════════════════════════════════════════════════════════════════

def _detect_encoding(path: str | Path) -> str:
    raw = open(path, "rb").read(4)
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16"
    if raw[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    return "utf-8"


def _detect_delimiter(line: str) -> str:
    return "\t" if line.count("\t") > line.count(",") else ","


def _find_column(headers: list[str], candidates: list[str]) -> int:
    """Return 0-based index of header best matching any candidate.

    Prefers EXACT match (case-insensitive, trimmed) over substring match. This
    avoids "Keyword" matching "Keyword status" when "Keyword" exists as its
    own column. Substring fallback only kicks in when no exact match found.
    """
    headers_lower = [(i, str(h or "").lower().strip()) for i, h in enumerate(headers)]
    candidates_lower = [c.lower().strip() for c in candidates]

    # 1st pass: exact match
    for i, h_low in headers_lower:
        if not h_low:
            continue
        if h_low in candidates_lower:
            return i

    # 2nd pass: substring match (original behavior, as fallback)
    for i, h_low in headers_lower:
        if not h_low:
            continue
        if any(c in h_low for c in candidates_lower):
            return i

    return -1


COL_ALIASES = {
    "campaign":   ["campaign", "campagne"],
    "ad_group":   ["ad group", "advertentiegroep"],
    "keyword":    ["keyword", "zoekwoord"],
    "match_type": ["match type", "type overeenkomst"],
    "status":     ["status"],
    "final_url":  ["final url", "uiteindelijke url"],
}


def read_keywords_csv(csv_path: str | Path) -> list[KeywordRow]:
    """Parse a Google Ads keywords export CSV. Returns list of KeywordRow."""
    encoding = _detect_encoding(csv_path)
    with open(csv_path, encoding=encoding) as f:
        lines = f.readlines()
    if not lines:
        return []

    # Find header row — first line with recognized column AND separator
    header_idx = 0
    for i, line in enumerate(lines[:8]):
        low = line.lower()
        has_sep = ("\t" in line) or (line.count(",") >= 2)
        if has_sep and any(
            any(a.lower() in low for a in aliases)
            for aliases in COL_ALIASES.values()
        ):
            header_idx = i
            break

    delimiter = _detect_delimiter(lines[header_idx])
    relevant = lines[header_idx:]
    reader = csv.reader(relevant, delimiter=delimiter)
    rows = []
    for row in reader:
        if not row or not any(c.strip() for c in row):
            continue
        if row and row[0].strip().lower().startswith("total"):
            continue
        rows.append(row)

    if len(rows) < 2:
        return []

    headers = rows[0]
    data = rows[1:]

    col_idx = {key: _find_column(headers, aliases) for key, aliases in COL_ALIASES.items()}
    if col_idx["keyword"] < 0:
        raise ValueError(f"Could not find 'Keyword' column. Headers: {headers}")

    out = []
    for row_i, row in enumerate(data, start=2):
        def get(key, default=""):
            idx = col_idx.get(key, -1)
            if idx >= 0 and idx < len(row):
                return str(row[idx]).strip()
            return default

        raw_kw = get("keyword")
        if not raw_kw:
            continue
        kw = _strip_decoration(raw_kw)
        norm = _normalize_keyword(raw_kw)
        match_type = _normalize_match_type(get("match_type"))

        out.append(KeywordRow(
            raw_keyword=raw_kw,
            keyword=kw,
            norm_keyword=norm,
            match_type=match_type,
            raw_match_type=get("match_type"),
            campaign=get("campaign"),
            ad_group=get("ad_group"),
            status=get("status") or "Active",
            final_url=get("final_url"),
            row_index=row_i,
        ))
    return out


# ═════════════════════════════════════════════════════════════════════════════
# Conflict detection
# ═════════════════════════════════════════════════════════════════════════════

def _is_serving(status: str) -> bool:
    """Whether a keyword is currently serving (eligible to show ads).

    Google Ads' effective Status value reflects parent campaign and ad-group
    pause state — if either is paused, the keyword's Status is "Paused" too.
    We also exclude 'Not eligible' (low-quality/disapproved keywords that
    don't actually serve). 'Limited' keywords still serve and are kept.
    """
    if not status:
        return True  # missing status — assume active
    s = status.lower().strip()
    if s in ("paused", "not eligible"):
        return False
    return True


def detect_conflicts(rows: list[KeywordRow]) -> list[Conflict]:
    """Run all four detectors. Returns deduplicated list of conflicts.

    Skips non-serving keywords (Paused, Not eligible). Google Ads' effective
    Status already reflects parent campaign / ad-group pause state, so a
    keyword inside a paused campaign or paused ad group is automatically
    excluded — no separate campaign-status / ad-group-status column needed.
    """
    # Filter out non-serving keywords
    active_rows = [r for r in rows if _is_serving(r.status)]
    conflicts: list[Conflict] = []

    # Helper: a unique "ad group instance" is the (campaign, ad_group) tuple,
    # because the same ad-group name can exist in different campaigns and they
    # are distinct serving units.
    def _ag_key(r: KeywordRow) -> tuple[str, str]:
        return (r.campaign, r.ad_group)

    # ── 1. EXACT_DUPLICATE: same norm_keyword + same match_type in 2+ (campaign, ad_group) pairs ─
    by_kw_mt: dict[tuple[str, str], list[KeywordRow]] = defaultdict(list)
    for r in active_rows:
        by_kw_mt[(r.norm_keyword, r.match_type)].append(r)
    for (norm, mt), group in by_kw_mt.items():
        ag_keys = {_ag_key(r) for r in group}
        if len(ag_keys) > 1:
            conflicts.append(_make_conflict("EXACT_DUPLICATE", group))

    # ── 2. MATCH_TYPE_OVERLAP: same norm_keyword, multiple match types, multiple ag_keys ─
    by_kw: dict[str, list[KeywordRow]] = defaultdict(list)
    for r in active_rows:
        by_kw[r.norm_keyword].append(r)
    for norm, group in by_kw.items():
        match_types = {r.match_type for r in group}
        ag_keys = {_ag_key(r) for r in group}
        # Need different match types AND different ad-group instances
        if len(match_types) > 1 and len(ag_keys) > 1:
            conflicts.append(_make_conflict("MATCH_TYPE_OVERLAP", group))

    # ── 3. PHRASE_CONTAINS: Broad keyword that's substring of other keywords in different ag_keys ─
    broad_rows = [r for r in active_rows if r.match_type in ("BROAD", "BROAD_MOD")]
    for broad in broad_rows:
        broad_norm = broad.norm_keyword
        if not broad_norm:
            continue
        broad_key = _ag_key(broad)
        # Find other keywords (any match type) where the broad keyword is a sub-phrase
        contained = []
        for r in active_rows:
            if r.row_index == broad.row_index:
                continue
            if _ag_key(r) == broad_key:
                continue  # same campaign+ad-group — intentional
            if r.norm_keyword == broad_norm:
                continue  # would be EXACT_DUPLICATE already
            # Word-boundary check — broad keyword must appear as whole-word sub-phrase
            pattern = r"\b" + re.escape(broad_norm) + r"\b"
            if re.search(pattern, r.norm_keyword):
                contained.append(r)
        if contained:
            group = [broad] + contained
            conflicts.append(_make_conflict("PHRASE_CONTAINS", group))

    # ── 4. CLOSE_VARIANT: same stem, different keyword text, different ag_keys ─
    by_stem: dict[str, list[KeywordRow]] = defaultdict(list)
    for r in active_rows:
        by_stem[_stem_lite(r.keyword)].append(r)
    for stem, group in by_stem.items():
        norms = {r.norm_keyword for r in group}
        ag_keys = {_ag_key(r) for r in group}
        if len(norms) > 1 and len(ag_keys) > 1:
            # Make sure not already fully covered as EXACT_DUPLICATE
            already_handled = any(
                len(by_kw[norm]) > 1 for norm in norms
            )
            if not already_handled:
                conflicts.append(_make_conflict("CLOSE_VARIANT", group))

    # Sort by severity, then by # keywords involved
    conflicts.sort(key=lambda c: (c.severity_rank, -len(c.keywords_involved)))
    return conflicts


def _make_conflict(ctype: str, rows: list[KeywordRow]) -> Conflict:
    campaigns = list({r.campaign for r in rows})
    geo = _is_geo_segmented(campaigns)
    all_paused = all(r.status.lower() == "paused" for r in rows)
    return Conflict(
        conflict_type=ctype,
        keywords_involved=rows,
        geo_segmented=geo,
        all_paused=all_paused,
    )


# ═════════════════════════════════════════════════════════════════════════════
# Excel output
# ═════════════════════════════════════════════════════════════════════════════

def build_excel(conflicts: list[Conflict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Overlap Issues"

    headers = [
        "Conflict type",
        "Keyword(s)",
        "Match types",
        "Ad groups",
        "Campaigns",
        "Probably geo-segmented?",
        "All paused?",
        "# instances",
    ]
    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font

    for row_idx, c in enumerate(conflicts, 2):
        kws = list({r.keyword for r in c.keywords_involved})
        mts = list({r.match_type for r in c.keywords_involved})
        ags = list({r.ad_group for r in c.keywords_involved})
        camps = list({r.campaign for r in c.keywords_involved})

        vals = [
            c.conflict_type.replace("_", " ").title(),
            " | ".join(sorted(kws)),
            ", ".join(sorted(mts)),
            " | ".join(sorted(ags)),
            " | ".join(sorted(camps)),
            "TRUE" if c.geo_segmented else "FALSE",
            "TRUE" if c.all_paused else "FALSE",
            len(c.keywords_involved),
        ]
        fill = SEVERITY_FILL.get(c.conflict_type, FILL_GREY)
        if c.geo_segmented or c.all_paused:
            fill = FILL_GREY  # de-emphasize likely-intentional overlaps
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    widths = [22, 40, 22, 40, 40, 20, 12, 12]
    for col_letter, width in zip("ABCDEFGH", widths):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"

    # ── Sheet 2: Suggested Pauses (Editor-importable) ───────────────────────
    pauses_ws = wb.create_sheet("Suggested Pauses")
    pause_headers = [
        "Campaign", "Ad group", "Keyword", "Match type", "Status",
        "Conflict type", "Reason",
    ]
    for col_idx, h in enumerate(pause_headers, 1):
        cell = pauses_ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font

    pause_row = 2
    seen_pauses: set[tuple[str, str, int]] = set()  # (campaign, ad_group, row_index)
    for c in conflicts:
        # Don't suggest pauses for already-intentional overlaps
        if c.geo_segmented or c.all_paused:
            continue
        # Don't suggest pauses for CLOSE_VARIANT (low severity, often legitimate plurals)
        if c.conflict_type == "CLOSE_VARIANT":
            continue
        # For each conflict, suggest pausing all but the first instance.
        # User decides which to actually pause — this is just a starting point.
        instances = c.keywords_involved
        for r in instances[1:]:  # keep first, pause rest
            key = (r.campaign, r.ad_group, r.row_index)
            if key in seen_pauses:
                continue  # already suggested via another conflict — dedup
            seen_pauses.add(key)
            pauses_ws.cell(row=pause_row, column=1, value=r.campaign)
            pauses_ws.cell(row=pause_row, column=2, value=r.ad_group)
            # Wrap keyword in match-type decoration for Editor paste
            kw_decorated = r.keyword
            if r.match_type == "EXACT":
                kw_decorated = f"[{r.keyword}]"
            elif r.match_type == "PHRASE":
                kw_decorated = f'"{r.keyword}"'
            pauses_ws.cell(row=pause_row, column=3, value=kw_decorated)
            pauses_ws.cell(row=pause_row, column=4, value=r.raw_match_type)
            pauses_ws.cell(row=pause_row, column=5, value="Paused")
            pauses_ws.cell(row=pause_row, column=6, value=c.conflict_type.replace("_", " ").title())
            reason = (
                f"Conflicts with {len(instances)-1} other instance(s) "
                f"in campaign(s): {', '.join(sorted({r2.campaign for r2 in instances if r2.row_index != r.row_index}))}"
            )
            pauses_ws.cell(row=pause_row, column=7, value=reason)
            pause_row += 1

    pause_widths = [28, 28, 32, 14, 10, 22, 50]
    for col_letter, width in zip("ABCDEFG", pause_widths):
        pauses_ws.column_dimensions[col_letter].width = width
    pauses_ws.freeze_panes = "A2"

    # ── Sheet 3: Legend ──────────────────────────────────────────────────────
    leg = wb.create_sheet("Legend")
    leg_data = [
        ["Conflict type", "Severity", "What it means", "Recommended action"],
        ["Exact duplicate",   "High",   "Same keyword + same match type in multiple ad groups",
         "Pause duplicates, keep best-performing instance"],
        ["Match type overlap", "Medium", "Same keyword text, different match types in multiple ad groups (Broad will absorb traffic from Exact)",
         "Decide which match type wins; pause the others"],
        ["Phrase contains",    "Medium", "Broad keyword X is contained in other keywords in different ad groups",
         "Either narrow Broad with negatives, or pause it"],
        ["Close variant",      "Low",    "Singular/plural / minor stem variations across ad groups",
         "Often safe to leave; review only if volume is high"],
        [],
        ["Annotation", "", "What it means"],
        ["Probably geo-segmented", "", "Campaign names suggest different country/region targeting → likely intentional, review only"],
        ["All paused",             "", "All instances of this conflict are already paused — informational"],
    ]
    fills_leg = [None, FILL_RED, FILL_YELLOW, FILL_YELLOW, FILL_GREEN, None, None, FILL_GREY, FILL_GREY]
    for r_idx, (row_data, fill) in enumerate(zip(leg_data, fills_leg), 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = leg.cell(row=r_idx, column=c_idx, value=val)
            if r_idx in (1, 7):
                cell.font = Font(bold=True)
            if fill and c_idx == 1:
                cell.fill = fill
    leg_widths = [22, 12, 60, 50]
    for col_letter, width in zip("ABCD", leg_widths):
        leg.column_dimensions[col_letter].width = width

    wb.save(output_path)


# ═════════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print("Usage: python detect_overlap.py <keywords_csv> [output_path]")
        sys.exit(1)

    csv_path = sys.argv[1]
    if not Path(csv_path).exists():
        print(f"ERROR: file not found: {csv_path}")
        sys.exit(1)

    if len(sys.argv) >= 3 and sys.argv[2]:
        output_path = sys.argv[2]
    else:
        base = os.path.splitext(os.path.basename(csv_path))[0]
        output_path = os.path.join(os.path.dirname(csv_path), f"keyword_overlap_{base}.xlsx")

    print(f"Reading: {csv_path}")
    rows = read_keywords_csv(csv_path)
    print(f"  Loaded {len(rows)} keywords")

    serving = sum(1 for r in rows if _is_serving(r.status))
    not_serving = len(rows) - serving
    print(f"  Serving: {serving}, Not serving (paused / not eligible / disapproved): {not_serving}")

    conflicts = detect_conflicts(rows)
    print(f"\nConflicts detected: {len(conflicts)}")
    by_type: dict[str, int] = defaultdict(int)
    for c in conflicts:
        by_type[c.conflict_type] += 1
    for ctype, count in sorted(by_type.items()):
        print(f"  {ctype:22}: {count}")

    geo_count = sum(1 for c in conflicts if c.geo_segmented)
    if geo_count:
        print(f"\n  Probably geo-segmented (likely intentional): {geo_count}")

    build_excel(conflicts, output_path)
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
